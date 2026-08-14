#!/usr/bin/env python3
"""
ProjectHub — Professional XLSX Report Generator
Menghasilkan laporan Excel report-ready (BUKAN dump database).
7 sheet: Summary, Milestones, Deliverables, Tasks, Issues&Risks, Progress History, Activity/Audit.

Usage:
  python excel_report.py <project_json> <output.xlsx>
Jika <project_json> tidak ada, dipakai SAMPLE_DATA (project P001).
"""
import sys, json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.pagebreak import Break

# ---------- palette ----------
BRAND      = "2F55D4"
BRAND_DK   = "1E3A8A"
LIGHT      = "EEF2FF"
ZEBRA      = "F5F7FF"
GREY       = "6B7280"
WHITE      = "FFFFFF"
G_GREEN    = "C6EFCE"; T_GREEN = "0B6B2E"
G_YELLOW   = "FFF2CC"; T_YELLOW= "8A6D00"
G_ORANGE   = "FCE4CC"; T_ORANGE= "9C4A00"
G_RED      = "FFC7CE"; T_RED   = "9C0006"
G_BLUE     = "DDEBFF"; T_BLUE  = "1E3A8A"

FONT = "Arial"
thin = Side(style="thin", color="D0D6E2")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def title_font(sz=18, color=BRAND, bold=True):   return Font(name=FONT, size=sz, bold=bold, color=color)
def hdr_font():                                  return Font(name=FONT, size=10, bold=True, color=WHITE)
def lbl_font(bold=True):                          return Font(name=FONT, size=10, bold=bold, color=BRAND_DK)
def body_font():                                 return Font(name=FONT, size=10, color="222222")
def small_font():                                return Font(name=FONT, size=8, italic=True, color=GREY)

def fill(hexc): return PatternFill("solid", fgColor=hexc)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT= Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_T=Alignment(horizontal="left", vertical="top", wrap_text=True)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def header_row(ws, row, headers, start_col=1, fillc=BRAND):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col+j, value=h)
        c.font = hdr_font(); c.fill = fill(fillc); c.alignment = CEN; c.border = BORDER
    ws.row_dimensions[row].height = 26

def write_row(ws, row, values, start_col=1, zebra=False, aligns=None):
    for j, v in enumerate(values):
        c = ws.cell(row=row, column=start_col+j, value=v)
        c.font = body_font(); c.border = BORDER
        c.alignment = (aligns[j] if aligns else LEFT)
        if zebra: c.fill = fill(ZEBRA)

def page_setup(ws, landscape=True, title_rows="1:6"):
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.oddHeader.center.text = "&\"Arial\"&10&KBBBBBBProjectHub — Project Status Report"
    ws.oddFooter.left.text  = "&\"Arial\"&8&K888888Confidential — Internal Use"
    ws.oddFooter.center.text= "&\"Arial\"&8&K888888Page &P of &N"
    ws.oddFooter.right.text = "&\"Arial\"&8&K888888Generated &D"

# health/status color map for conditional formatting text values
def apply_status_cf(ws, col_letter, first, last):
    rng = f"{col_letter}{first}:{col_letter}{last}"
    rules = [
        ("done",        G_GREEN,  T_GREEN),
        ("on_track",    G_GREEN,  T_GREEN),
        ("On Track",    G_GREEN,  T_GREEN),
        ("Done",        G_GREEN,  T_GREEN),
        ("in_progress", G_BLUE,   T_BLUE),
        ("In Progress", G_BLUE,   T_BLUE),
        ("watch",       G_YELLOW, T_YELLOW),
        ("Watch",       G_YELLOW, T_YELLOW),
        ("at_risk",     G_ORANGE, T_ORANGE),
        ("At Risk",     G_ORANGE, T_ORANGE),
        ("critical",    G_RED,    T_RED),
        ("Critical",    G_RED,    T_RED),
        ("high",        G_ORANGE, T_ORANGE),
        ("High",        G_ORANGE, T_ORANGE),
        ("medium",      G_YELLOW, T_YELLOW),
        ("Medium",      G_YELLOW, T_YELLOW),
        ("open",        G_ORANGE, T_ORANGE),
        ("Open",        G_ORANGE, T_ORANGE),
    ]
    for text, bg, fg in rules:
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="containsText", formula=[f'"{text}"'],  # placeholder; use text rule below
            stopIfTrue=False, fill=fill(bg), font=Font(name=FONT, size=10, bold=True, color=fg)))

# openpyxl CellIsRule doesn't do containsText; use text rule via FormulaRule
def apply_text_cf(ws, col_letter, first, last):
    rng = f"{col_letter}{first}:{col_letter}{last}"
    mapping = [
        (["Done","done","On Track","on_track","Approved","approved","Resolved","resolved"], G_GREEN, T_GREEN),
        (["In Progress","in_progress","Monitoring","monitoring"], G_BLUE, T_BLUE),
        (["Watch","watch","Medium","medium"], G_YELLOW, T_YELLOW),
        (["At Risk","at_risk","High","high","Open","open","Delayed","delayed"], G_ORANGE, T_ORANGE),
        (["Critical","critical","Overdue","Rejected","rejected"], G_RED, T_RED),
    ]
    for texts, bg, fg in mapping:
        for t in texts:
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'ISNUMBER(SEARCH("{t}",{col_letter}{first}))'],
                stopIfTrue=False, fill=fill(bg),
                font=Font(name=FONT, size=10, bold=True, color=fg)))

# =====================================================================
def build(data, out):
    P = data["project"]
    period = data.get("period", "April – Mei 2026")
    gen_ts = datetime.datetime.now().strftime("%d %b %Y %H:%M")

    wb = Workbook()

    # ---------------- SHEET 1: SUMMARY ----------------
    ws = wb.active; ws.title = "Summary"
    set_widths(ws, [3, 22, 26, 22, 26])
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:E2"); ws["B2"] = "PROJECT STATUS REPORT"; ws["B2"].font = title_font(20)
    ws.merge_cells("B3:E3"); ws["B3"] = P["name"]; ws["B3"].font = Font(name=FONT, size=13, color=BRAND_DK, bold=True)
    ws.merge_cells("B4:E4"); ws["B4"] = f"Period: {period}   ·   Generated: {gen_ts}"; ws["B4"].font = small_font()

    # info block
    info = [
        ("Project Code", P["code"], "Owner", P["owner"]),
        ("Category", P.get("category","-"), "Sponsor", P.get("sponsor","-")),
        ("Start Date", P.get("start","-"), "Target Date", P.get("target","-")),
        ("Status", P.get("status","-"), "Priority", P.get("priority","-")),
    ]
    r = 6
    for a,b,c,d in info:
        ws.cell(r,2,a).font = lbl_font(); ws.cell(r,2).fill=fill(LIGHT); ws.cell(r,2).border=BORDER; ws.cell(r,2).alignment=LEFT
        ws.cell(r,3,b).font = body_font(); ws.cell(r,3).border=BORDER; ws.cell(r,3).alignment=LEFT
        ws.cell(r,4,c).font = lbl_font(); ws.cell(r,4).fill=fill(LIGHT); ws.cell(r,4).border=BORDER; ws.cell(r,4).alignment=LEFT
        ws.cell(r,5,d).font = body_font(); ws.cell(r,5).border=BORDER; ws.cell(r,5).alignment=LEFT
        r += 1

    # KPI band
    r += 1
    ws.cell(r,2,"KEY METRICS").font = Font(name=FONT, size=11, bold=True, color=BRAND); r += 1
    kpi = P["kpi"]
    kpis = [("Overall Progress", f'{kpi["progress"]}%'), ("Health", kpi["health"]),
            ("Open Issues", kpi["open_issues"]), ("High/Critical Risks", kpi["high_risks"]),
            ("Overdue Tasks", kpi["overdue"]), ("Milestones Done", kpi["ms_done"])]
    col = 2
    kpi_row = r
    for label, val in kpis:
        ws.cell(kpi_row, col, label).font = Font(name=FONT, size=9, color=GREY)
        ws.cell(kpi_row, col).alignment = CEN; ws.cell(kpi_row, col).fill=fill(LIGHT); ws.cell(kpi_row,col).border=BORDER
        vc = ws.cell(kpi_row+1, col, val)
        vc.font = Font(name=FONT, size=14, bold=True, color=BRAND_DK); vc.alignment = CEN; vc.border=BORDER
        col += 1
        if col > 5:  # wrap after 4 cols (B..E)
            col = 2; kpi_row += 2
    # health reason
    hr = kpi_row + 2
    ws.merge_cells(start_row=hr, start_column=2, end_row=hr, end_column=5)
    hc = ws.cell(hr,2, f'Health: {kpi["health"]} — {kpi.get("health_reason","-")}')
    hc.font = Font(name=FONT, size=10, bold=True, color=T_ORANGE); hc.alignment = LEFT_T
    ws.freeze_panes = "A6"
    page_setup(ws, landscape=False)

    # ---------------- SHEET 2: MILESTONES ----------------
    ws = wb.create_sheet("Milestones")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1"); ws["A1"]="MILESTONES"; ws["A1"].font=title_font(14)
    ws.merge_cells("A2:G2"); ws["A2"]=f"{P['name']} · {period}"; ws["A2"].font=small_font()
    hdrs = ["Milestone","Owner","Target Date","Progress","Status"]
    set_widths(ws, [34,16,14,12,16])
    header_row(ws, 4, hdrs)
    r = 5
    for m in data["milestones"]:
        write_row(ws, r, [m["name"], m["owner"], m["target"], m["progress"]/100, m["status"]],
                  aligns=[LEFT,CEN,CEN,CEN,CEN], zebra=(r%2==0))
        ws.cell(r,4).number_format = "0%"
        r += 1
    apply_text_cf(ws, "E", 5, r-1)
    ws.auto_filter.ref = f"A4:E{r-1}"
    ws.freeze_panes = "A5"
    page_setup(ws)

    # ---------------- SHEET 3: DELIVERABLES ----------------
    ws = wb.create_sheet("Deliverables")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1"); ws["A1"]="DELIVERABLES"; ws["A1"].font=title_font(14)
    ws.merge_cells("A2:F2"); ws["A2"]=f"{P['name']} · {period}"; ws["A2"].font=small_font()
    hdrs = ["Deliverable","Milestone","Owner","Target Date","Progress","Status"]
    set_widths(ws, [38,22,16,14,12,16])
    header_row(ws, 4, hdrs)
    r = 5
    for d in data["deliverables"]:
        write_row(ws, r, [d["name"], d["milestone"], d["owner"], d["target"], d["progress"]/100, d["status"]],
                  aligns=[LEFT,LEFT,CEN,CEN,CEN,CEN], zebra=(r%2==0))
        ws.cell(r,5).number_format="0%"
        r += 1
    apply_text_cf(ws, "F", 5, r-1)
    ws.auto_filter.ref = f"A4:F{r-1}"
    ws.freeze_panes = "A5"; page_setup(ws)

    # ---------------- SHEET 4: TASKS ----------------
    ws = wb.create_sheet("Tasks")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:I1"); ws["A1"]="TASKS"; ws["A1"].font=title_font(14)
    ws.merge_cells("A2:I2"); ws["A2"]=f"{P['name']} · {period}"; ws["A2"].font=small_font()
    hdrs = ["Task Code","Task","PIC","Status","Progress","Health","Due Date","Constraint","Next Action"]
    set_widths(ws, [12,34,14,14,10,12,12,26,26])
    header_row(ws, 4, hdrs)
    r = 5
    for t in data["tasks"]:
        write_row(ws, r, [t["code"], t["title"], t["pic"], t["status"], t["progress"]/100,
                          t["health"], t["due"], t.get("constraint","") or "", t.get("next_action","") or ""],
                  aligns=[CEN,LEFT,CEN,CEN,CEN,CEN,CEN,LEFT,LEFT], zebra=(r%2==0))
        ws.cell(r,5).number_format="0%"
        r += 1
    apply_text_cf(ws, "D", 5, r-1)
    apply_text_cf(ws, "F", 5, r-1)
    ws.auto_filter.ref = f"A4:I{r-1}"
    ws.freeze_panes = "C5"; page_setup(ws)

    # ---------------- SHEET 5: ISSUES & RISKS ----------------
    ws = wb.create_sheet("Issues & Risks")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:I1"); ws["A1"]="ISSUES & RISKS"; ws["A1"].font=title_font(14)
    ws.merge_cells("A2:I2"); ws["A2"]=f"{P['name']} · {period}"; ws["A2"].font=small_font()
    hdrs = ["Type","Title","Severity","Impact","Probability","Owner","Status","Target Date","Next Action"]
    set_widths(ws, [10,32,12,24,12,14,14,13,26])
    header_row(ws, 4, hdrs)
    r = 5
    for x in data["issues_risks"]:
        write_row(ws, r, [x["type"], x["title"], x["severity"], x.get("impact",""), x.get("probability",""),
                          x["owner"], x["status"], x.get("target",""), x.get("next_action","")],
                  aligns=[CEN,LEFT,CEN,LEFT,CEN,CEN,CEN,CEN,LEFT], zebra=(r%2==0))
        r += 1
    apply_text_cf(ws, "C", 5, r-1)
    apply_text_cf(ws, "G", 5, r-1)
    ws.auto_filter.ref = f"A4:I{r-1}"
    ws.freeze_panes = "A5"; page_setup(ws)

    # ---------------- SHEET 6: PROGRESS HISTORY ----------------
    ws = wb.create_sheet("Progress History")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1"); ws["A1"]="PROGRESS HISTORY"; ws["A1"].font=title_font(14)
    ws.merge_cells("A2:G2"); ws["A2"]=f"{P['name']} · {period}"; ws["A2"].font=small_font()
    hdrs = ["Date","Entity","Previous","New","Delta","Updated By","Comment"]
    set_widths(ws, [18,16,10,10,10,16,30])
    header_row(ws, 4, hdrs)
    r = 5
    for h in data["progress_history"]:
        write_row(ws, r, [h["date"], h["entity"], h["prev"]/100, h["new"]/100, h["delta"]/100, h["by"], h.get("comment","")],
                  aligns=[CEN,CEN,CEN,CEN,CEN,CEN,LEFT], zebra=(r%2==0))
        for cc in (3,4,5): ws.cell(r,cc).number_format="0%"
        # delta color
        dc = ws.cell(r,5)
        dc.font = Font(name=FONT, size=10, bold=True, color=(T_GREEN if h["delta"]>=0 else T_RED))
        r += 1
    ws.auto_filter.ref = f"A4:G{r-1}"
    ws.freeze_panes = "A5"; page_setup(ws)

    # ---------------- SHEET 7: ACTIVITY / AUDIT ----------------
    ws = wb.create_sheet("Activity & Audit")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1"); ws["A1"]="ACTIVITY & AUDIT TRAIL"; ws["A1"].font=title_font(14)
    ws.merge_cells("A2:G2"); ws["A2"]=f"{P['name']} · {period}"; ws["A2"].font=small_font()
    hdrs = ["Date","User","Action","Entity","Field","Previous Value","New Value"]
    set_widths(ws, [18,16,16,16,14,20,20])
    header_row(ws, 4, hdrs)
    r = 5
    for a in data["audit"]:
        write_row(ws, r, [a["date"], a["user"], a["action"], a["entity"], a.get("field",""),
                          a.get("prev",""), a.get("new","")],
                  aligns=[CEN,CEN,CEN,CEN,CEN,LEFT,LEFT], zebra=(r%2==0))
        r += 1
    ws.auto_filter.ref = f"A4:G{r-1}"
    ws.freeze_panes = "A5"; page_setup(ws)

    wb.save(out)
    return out

# =====================================================================
SAMPLE_DATA = {
    "period": "April – Mei 2026",
    "project": {
        "code":"P001","name":"Monitoring PM Excel","category":"Trial","owner":"Hanif Mu'taz",
        "sponsor":"Management","start":"01 Apr 2026","target":"06 Mei 2026","status":"Execution","priority":"High",
        "kpi":{"progress":72,"health":"At Risk","open_issues":2,"high_risks":1,"overdue":0,"ms_done":1,
               "health_reason":"1 high-impact issue · schedule variance 18% (waktu 90% vs progress 72%)"}
    },
    "milestones":[
        {"name":"M1 - Planning","owner":"Hanif","target":"15 Apr 2026","progress":100,"status":"Done"},
        {"name":"M2 - Development","owner":"Hanif","target":"25 Apr 2026","progress":80,"status":"In Progress"},
        {"name":"M3 - Testing & Sign-off","owner":"Hanif","target":"06 Mei 2026","progress":30,"status":"In Progress"},
    ],
    "deliverables":[
        {"name":"Requirement & Template Excel","milestone":"M1 - Planning","owner":"Hanif","target":"15 Apr 2026","progress":100,"status":"Done"},
        {"name":"VBA Report Engine","milestone":"M2 - Development","owner":"Hanif","target":"25 Apr 2026","progress":75,"status":"In Progress"},
        {"name":"Validated PM Monitoring V1","milestone":"M3 - Testing & Sign-off","owner":"Hanif","target":"06 Mei 2026","progress":30,"status":"In Progress"},
    ],
    "tasks":[
        {"code":"P001.1","title":"Analisa kebutuhan data PM","pic":"Hanif","status":"Done","progress":100,"health":"On Track","due":"10 Apr","constraint":"","next_action":""},
        {"code":"P001.2","title":"Desain template Excel","pic":"Hanif","status":"Done","progress":100,"health":"On Track","due":"15 Apr","constraint":"","next_action":""},
        {"code":"P001.3","title":"Develop VBA import & parsing","pic":"Hanif","status":"Done","progress":100,"health":"On Track","due":"20 Apr","constraint":"","next_action":""},
        {"code":"P001.4","title":"Develop VBA laporan per-short","pic":"Hanif","status":"Done","progress":100,"health":"On Track","due":"25 Apr","constraint":"","next_action":""},
        {"code":"P001.5","title":"Develop VBA laporan monthly & weekly","pic":"Hanif","status":"In Progress","progress":50,"health":"Watch","due":"06 Mei","constraint":"","next_action":"Lanjut coding aggregate"},
        {"code":"P001.6","title":"Meeting request feature baru","pic":"Hanif","status":"In Progress","progress":40,"health":"At Risk","due":"06 Mei","constraint":"","next_action":"Siapkan deck request"},
        {"code":"P001.7","title":"Trial & validasi data real","pic":"Hanif","status":"In Progress","progress":30,"health":"At Risk","due":"06 Mei","constraint":"Data belum matching manual","next_action":"Cek hasil vs data manual"},
        {"code":"P001.8","title":"Review & sign-off ke direktur","pic":"Hanif","status":"In Progress","progress":20,"health":"At Risk","due":"06 Mei","constraint":"","next_action":"Presentasi hasil trial"},
    ],
    "issues_risks":[
        {"type":"Issue","title":"Data real belum matching hasil manual","severity":"High","impact":"Validasi tertunda, sign-off berisiko mundur","probability":"","owner":"Hanif","status":"Open","target":"04 Mei","next_action":"Cek formula parsing vs manual"},
        {"type":"Issue","title":"Deck request feature belum siap","severity":"Medium","impact":"Meeting bisa mundur","probability":"","owner":"Hanif","status":"Open","target":"05 Mei","next_action":"Selesaikan deck"},
        {"type":"Risk","title":"Sign-off direktur mundur","severity":"High","impact":"Timeline mundur","probability":"High","owner":"Hanif","status":"Mitigating","target":"06 Mei","next_action":"Buffer 3 hari"},
    ],
    "progress_history":[
        {"date":"01 Mei 09:00","entity":"P001.7","prev":0,"new":10,"delta":10,"by":"Hanif","comment":"Mulai trial"},
        {"date":"02 Mei 10:00","entity":"P001.7","prev":10,"new":20,"delta":10,"by":"Hanif","comment":"Setup data real"},
        {"date":"03 Mei 14:00","entity":"P001.7","prev":20,"new":30,"delta":10,"by":"Hanif","comment":"Cek batch pertama"},
        {"date":"20 Apr 11:00","entity":"P001.5","prev":0,"new":25,"delta":25,"by":"Hanif","comment":"Skeleton aggregate"},
        {"date":"28 Apr 16:00","entity":"P001.5","prev":25,"new":50,"delta":25,"by":"Hanif","comment":"Monthly report jalan"},
    ],
    "audit":[
        {"date":"03 Mei 14:00","user":"Hanif","action":"status_change","entity":"P001.7","field":"progress","prev":"20","new":"30"},
        {"date":"01 Mei 08:30","user":"Hanif","action":"created","entity":"Issue","field":"-","prev":"-","new":"Data belum matching"},
        {"date":"28 Apr 16:00","user":"Hanif","action":"status_change","entity":"P001.5","field":"progress","prev":"25","new":"50"},
        {"date":"25 Apr 17:10","user":"Hanif","action":"completed","entity":"P001.4","field":"status","prev":"in_progress","new":"done"},
    ],
}

if __name__ == "__main__":
    out = sys.argv[2] if len(sys.argv) > 2 else "ProjectHub_Report.xlsx"
    if len(sys.argv) > 1 and sys.argv[1] not in ("-", "sample"):
        with open(sys.argv[1]) as f: data = json.load(f)
    else:
        data = SAMPLE_DATA
    build(data, out)
    print("Report written:", out)
